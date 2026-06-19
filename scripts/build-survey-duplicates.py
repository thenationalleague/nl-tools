"""Commercial Benchmarking — surface duplicate survey submissions.

The raw SurveyMonkey export had several clubs submitting more than once. When
the dataset was cleaned, ONE submission per club was kept. This script lays the
competing submissions side by side so the choice can be explained to the
commercial team: each duplicated club's submissions become a block of rows with
the key answers, a completeness score, and — if the cleaned RTDB export is
supplied — a marker showing which submission matches the figures that were kept.

Usage:
    python scripts/build-survey-duplicates.py <survey.xlsx> [clubs.json] [out.xlsx]

  survey.xlsx  raw SurveyMonkey export (two header rows; club picked via a
               one-column-per-club matrix between 'Club name' and 'Division').
  clubs.json   optional RTDB export of app-data/ops-commercial-benchmarking/clubs
               (or its parent) — used only to mark the kept submission.

NOTHING here is committed; the output names clubs and people. Treat as confidential.
"""
import sys, json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SURVEY = sys.argv[1] if len(sys.argv) > 1 else 'survey.xlsx'
CLUBS = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2].endswith('.json') else None
OUT = next((a for a in sys.argv[2:] if a.endswith('.xlsx')), 'commercial-benchmarking-duplicates.xlsx')

CLUB_BLOCK = range(10, 82)   # 'Club name' (10) + one column per club option, up to 'Division'(82)

# (header, column index in the raw export) — single-answer columns only
FIELDS = [
    ('Front shirt sponsor', 89), ('FS deal length', 108), ('FS income (£)', 109),
    ('Back shirt sponsor', 111), ('BS deal length', 130), ('BS income (£)', 131),
    ('Sleeve sponsor', 133), ('SL deal length', 152), ('SL income (£)', 153),
    ('Stand 1', 155), ('Stand 1 £', 175), ('Stand 2', 176), ('Stand 2 £', 196),
    ('Stand 3', 197), ('Stand 3 £', 217), ('Stand 4', 218), ('Stand 4 £', 238),
    ('TV board £', 240), ('Non-TV board £', 242),
    ('Programme?', 243), ('Programme format', 245), ('Full-page advert £', 248),
    ('Top matchday ticket £', 249), ('Top season ticket £', 250),
    ('Top matchday hosp £', 251), ('Top seasonal hosp £', 260),
    ('Can email supporters?', 270), ('Can email partners?', 272),
    ('Email database', 276), ('Opted-in', 277),
]
# fields used to decide which submission was KEPT (compared to the cleaned record)
MATCH = {'fsSponsor': 89, 'frontShirt': 109, 'msTicket': 249, 'seasonTicket': 250,
         'mdHosp': 251, 'emailDb': 276, 'backShirt': 131, 'sleeve': 153}

GREEN = PatternFill('solid', fgColor='C9E5C2')
AMBER = PatternFill('solid', fgColor='FFE9A8')
GREY  = PatternFill('solid', fgColor='EDEDED')
HEAD  = PatternFill('solid', fgColor='1B2A4A')
HEADF = Font(color='FFFFFF', bold=True)
BOLD  = Font(bold=True)


def norm(s):
    return ' '.join(str(s if s is not None else '').strip().lower().split())


def numish(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v if v is not None else '').replace(',', '').replace('£', '').strip()
    try:
        return float(s)
    except ValueError:
        return None


def clubof(r):
    for ci in CLUB_BLOCK:
        if r[ci] not in (None, ''):
            return str(r[ci]).strip()
    return None


def load_final(path):
    if not path:
        return {}
    with open(path, encoding='utf-8') as f:
        d = json.load(f)
    if isinstance(d, dict) and isinstance(d.get('clubs'), dict):
        d = d['clubs']
    out = {}
    for k, v in d.items():
        if isinstance(v, dict):
            out[norm(v.get('club', k))] = v
    return out


def mval(rec, key):
    m = (rec.get('metrics') or {}).get(key)
    return m.get('value') if isinstance(m, dict) else m


def main():
    wb_in = openpyxl.load_workbook(SURVEY, data_only=True)
    ws_in = wb_in.active
    data = list(ws_in.iter_rows(values_only=True))[2:]   # skip the two header rows
    final = load_final(CLUBS)

    # group submissions by club
    groups = {}
    for r in data:
        c = clubof(r)
        if not c:
            continue
        groups.setdefault(norm(c), {'name': c, 'rows': []})['rows'].append(r)
    dups = {k: g for k, g in groups.items() if len(g['rows']) > 1}

    def completeness(r):
        return sum(1 for _h, ci in FIELDS if r[ci] not in (None, ''))

    def match_score(r, rec):
        if not rec:
            return -1
        s = 0
        for key, ci in MATCH.items():
            fv, rv = r[ci], mval(rec, key)
            if rv is None or fv in (None, ''):
                continue
            a, b = numish(fv), numish(rv)
            if a is not None and b is not None:
                if abs(a - b) < 0.5:
                    s += 1
            elif norm(fv) and norm(fv) == norm(rv):
                s += 1
        return s

    wb = Workbook()
    ws = wb.active
    ws.title = 'Duplicate submissions'
    cols = ['Club', 'Submission', 'Kept?', 'Completeness', 'Respondent ID', 'Submitted'] + [h for h, _ in FIELDS]
    ws.append(cols)

    n_clubs = 0
    for key in sorted(dups, key=lambda k: (-len(dups[k]['rows']), dups[k]['name'])):
        g = dups[key]
        rec = final.get(key)
        n_clubs += 1
        rws = g['rows']
        scores = [match_score(r, rec) for r in rws]
        best = max(scores) if scores else -1
        kept_idx = scores.index(best) if best > 0 else -1
        for i, r in enumerate(rws):
            kept = 'KEPT' if i == kept_idx else ('matches' if scores[i] == best and best > 0 else '')
            line = [g['name'], '#%d of %d' % (i + 1, len(rws)), kept, completeness(r),
                    r[0], r[3]] + [r[ci] for _h, ci in FIELDS]
            ws.append(line)
            rr = ws.max_row
            fill = GREEN if i == kept_idx else (GREY if (i % 2) else None)
            if fill:
                for ci in range(1, len(cols) + 1):
                    ws.cell(rr, ci).fill = fill
            if i == kept_idx:
                ws.cell(rr, 1).font = BOLD
        ws.append([])  # blank row between clubs

    # styling
    for ci in range(1, ws.max_column + 1):
        cell = ws.cell(1, ci)
        cell.fill = HEAD; cell.font = HEADF
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.freeze_panes = 'B2'
    for ci in range(1, ws.max_column + 1):
        w = 12
        for ri in range(1, ws.max_row + 1):
            v = ws.cell(ri, ci).value
            if v is not None:
                w = max(w, min(34, len(str(v)) + 2))
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.column_dimensions['A'].width = 22

    wb.save(OUT)
    print('Wrote %s' % OUT)
    print('  %d clubs submitted more than once (%d submissions total across them)'
          % (n_clubs, sum(len(g['rows']) for g in dups.values())))
    if final:
        marked = sum(1 for k in dups if final.get(k) and max(
            (match_score(r, final.get(k)) for r in dups[k]['rows']), default=0) > 0)
        print('  kept submission identified for %d/%d via the cleaned export' % (marked, n_clubs))
    else:
        print('  (no clubs.json supplied — "Kept?" left blank; pass the RTDB export to mark it)')


if __name__ == '__main__':
    main()
