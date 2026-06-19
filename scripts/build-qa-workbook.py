"""Commercial Benchmarking — build a QA workbook for the commercial team.

Reads a JSON export of the benchmarking club rows and writes a tidy .xlsx that
flags the things worth a manual look before following up with clubs:

  * 'All clubs'  — one row per club, every submitted field. Big-outlier cells
                   are highlighted amber; rows in a suspected duplicate set are
                   tinted red. Status / Outlier notes / Duplicate-group columns.
  * 'Outliers'   — one row per flagged value (club, field, value, league median,
                   how many times the median it is), largest first.
  * 'Duplicates' — suspected duplicate groups, every member repeated together so
                   the team can compare them side by side ("include again").

The input JSON can be any of:
  * the RTDB node  app-data/ops-commercial-benchmarking  (has a `clubs` child),
  * just its `clubs` child  { "<club>": {...}, ... },
  * the import file produced by build-benchmarks.py (also has `clubs`).
Export it from the Firebase console: open the node, ⋮ → Export JSON.

Usage:
    python scripts/build-qa-workbook.py <data.json> [out.xlsx] [survey.xlsx]

If the raw SurveyMonkey export is passed as a third argument, each duplicated
club's IGNORED submissions are added beneath its kept row, struck through, so
it's visible which returns were superseded.

NOTHING here is committed — the workbook contains club-identifiable figures.
Treat the output as confidential.
"""
import sys, json, re, datetime
import statistics as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_args = sys.argv[1:]
_json = [a for a in _args if a.endswith('.json')]
_xlsx = [a for a in _args if a.endswith('.xlsx')]
SRC = _json[0] if _json else 'commercial-benchmarking-rtdb-import.json'
OUT = _xlsx[0] if _xlsx else 'commercial-benchmarking-QA.xlsx'
SURVEY = _xlsx[1] if len(_xlsx) > 1 else None

# (metric key, column header, numeric?, flag-outliers?)
METRICS = [
    ('msTicket',     'Top GA matchday ticket (£)',   True,  True),
    ('seasonTicket', 'Top GA season ticket (£)',     True,  True),
    ('frontShirt',   'Front shirt income (£)',       True,  True),
    ('frontTerm',    'Front shirt term (yrs)',       True,  True),
    ('backShirt',    'Back shirt income (£)',        True,  True),
    ('backTerm',     'Back shirt term (yrs)',        True,  True),
    ('sleeve',       'Sleeve income (£)',            True,  True),
    ('sleeveTerm',   'Sleeve term (yrs)',            True,  True),
    ('standCount',   'Stand sponsors (n)',           True,  False),
    ('standTotal',   'Stand total (£)',              True,  True),
    ('standAvg',     'Stand avg per stand (£)',      True,  True),
    ('tvBoard',      'TV-facing board (£)',          True,  True),
    ('nonTvBoard',   'Non-TV board (£)',             True,  True),
    ('mdHosp',       'Matchday hospitality (£)',     True,  True),
    ('seasonHosp',   'Seasonal hospitality (£)',     True,  True),
    ('progAd',       'Programme advert (£)',         True,  True),
    ('emailDb',      'Email database (n)',           True,  True),
    ('optedIn',      'Opted-in partner emails (n)',  True,  True),
]
SPON = [
    ('fsSponsor', 'Front shirt sponsor'), ('fsSector', 'Front shirt sector'),
    ('bsSponsor', 'Back shirt sponsor'),  ('bsSector', 'Back shirt sector'),
    ('slSponsor', 'Sleeve sponsor'),      ('slSector', 'Sleeve sector'),
]
CHIPS = [
    ('progFormat', 'Programme format'), ('rollingFront', 'Front rolling?'),
    ('rollingBack', 'Back rolling?'),   ('rollingSleeve', 'Sleeve rolling?'),
    ('emailSupporters', 'Can email supporters?'), ('emailPartners', 'Can email partners?'),
]
# core figures used to detect copy/paste duplicate returns
DUP_FIELDS = ['frontShirt', 'backShirt', 'sleeve', 'msTicket', 'seasonTicket',
              'mdHosp', 'seasonHosp', 'progAd', 'emailDb']
HDR = dict((k, h) for k, h, _n, _f in METRICS)

AMBER = PatternFill('solid', fgColor='FFE9A8')
RED   = PatternFill('solid', fgColor='F8C9C4')
HEAD  = PatternFill('solid', fgColor='1B2A4A')
HEADF = Font(color='FFFFFF', bold=True)
THIN  = Border(*([Side(style='thin', color='DDDDDD')] * 4))


def load_clubs(path):
    with open(path, encoding='utf-8') as f:
        d = json.load(f)
    if isinstance(d, dict) and isinstance(d.get('clubs'), dict):
        d = d['clubs']
    clubs = []
    for k, v in d.items():
        if isinstance(v, dict):
            v = dict(v)
            v.setdefault('club', k)
            clubs.append(v)
    return clubs


def mval(c, key):
    m = (c.get('metrics') or {}).get(key)
    return m.get('value') if isinstance(m, dict) else m


def isnum(x):
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def money(v):
    if not isnum(v):
        return v if v not in (None, '') else ''
    return round(v, 2) if v % 1 else int(v)


PLACEHOLDER = {'', '0', '-', '–', '—', 'n/a', 'na', 'n.a.', 'none', 'nil', 'tbc', 'tbd', 'vacant'}


def is_placeholder(s):
    return str(s if s is not None else '').strip().lower() in PLACEHOLDER


def norm(s):
    return ' '.join(str(s if s is not None else '').strip().lower().split())


def pos(x):
    return isnum(x) and x > 0


def name_slots(c):
    """All sponsor-bearing slots for a club: the three shirt slots + stands."""
    slots = [('Front shirt', c.get('fsSponsor'), None),
             ('Back shirt',  c.get('bsSponsor'), None),
             ('Sleeve',      c.get('slSponsor'), None)]
    for i, sd in enumerate(c.get('stands') or [], 1):
        if isinstance(sd, dict):
            slots.append(('Stand %d' % i, sd.get('name'), sd.get('income')))
    return slots


def within_dups(c):
    """Same sponsor name appearing in >1 slot of one club (case-insensitive).
    Returns {normalised name: [(slot, raw, income), ...]}."""
    seen = {}
    for slot, raw, inc in name_slots(c):
        if is_placeholder(raw):
            continue
        seen.setdefault(norm(raw), []).append((slot, raw, inc))
    return {n: v for n, v in seen.items() if len(v) > 1}


def placeholder_stands(c):
    """Stand slots that are placeholders / vacant yet still counted."""
    out = []
    for i, sd in enumerate(c.get('stands') or [], 1):
        if isinstance(sd, dict) and is_placeholder(sd.get('name')) and not pos(sd.get('income')):
            out.append(i)
    return out


def main():
    clubs = load_clubs(SRC)
    if not clubs:
        print('No clubs found in %s' % SRC); sys.exit(1)
    DIV = {'National': 0, 'North': 1, 'South': 2}
    clubs.sort(key=lambda c: (DIV.get(c.get('division'), 9), str(c.get('club', ''))))

    # ---- outlier fences per numeric metric (robust: IQR*3 and 6x median) ----
    fences = {}
    for key, _h, _isnum, flag in METRICS:
        if not (_isnum and flag):
            continue
        vals = sorted(v for v in (mval(c, key) for c in clubs) if isnum(v))
        if len(vals) < 4:
            continue
        n = len(vals)
        q1 = vals[int(round(0.25 * (n - 1)))]
        q3 = vals[int(round(0.75 * (n - 1)))]
        iqr = q3 - q1
        med = st.median(vals)
        hi = max(q3 + 3 * iqr, med * 6) if med > 0 else q3 + 3 * iqr
        lo = q1 - 3 * iqr
        fences[key] = (lo, hi, med)

    def outliers_for(c):
        out = []
        for key, _h, _num, _flag in METRICS:
            if key not in fences:
                continue
            v = mval(c, key)
            if not isnum(v):
                continue
            lo, hi, med = fences[key]
            if v > hi or v < lo:
                out.append((key, v, med, (v / med) if med else None))
        return out

    # ---- data-quality flags per club ----
    TERM_ROLL = {'frontTerm': ('rollingFront', 'Front'), 'backTerm': ('rollingBack', 'Back'),
                 'sleeveTerm': ('rollingSleeve', 'Sleeve')}
    LONG_TERM = 10  # yrs — a deal this long usually means "rolling", not a fixed term

    def quality_flags(c):
        flags = []
        # NB: the same sponsor appearing in more than one slot is expected (a
        # sponsor can buy several assets), so it is deliberately NOT flagged.
        # blank / "0" / vacant stands that are still counted in the stand total
        ph = placeholder_stands(c)
        if ph:
            flags.append(('Placeholder stand', 'Stand %s blank/"0"/vacant but counted (stand count = %s)'
                          % (', '.join(map(str, ph)), money(mval(c, 'standCount')))))
        # implausibly long deal lengths — usually a rolling deal entered as a number
        for tk, (rk, lab) in TERM_ROLL.items():
            v = mval(c, tk)
            if isnum(v) and v >= LONG_TERM:
                roll = str((c.get('chips') or {}).get(rk, '')).strip()
                extra = ' but marked NOT rolling' if roll.lower() == 'no' else (' (rolling = %s)' % roll if roll else '')
                flags.append(('Long deal length', '%s-shirt term = %s yrs%s — confirm if rolling' % (lab, money(v), extra)))
        # two sectors packed into one stand field
        for i, sd in enumerate(c.get('stands') or [], 1):
            if isinstance(sd, dict) and '|' in str(sd.get('sector', '')):
                flags.append(('Malformed sector', 'Stand %d sector holds two values: "%s"' % (i, sd.get('sector'))))
        return flags

    def has_data(c):
        return any(k != 'standCount' and isnum(mval(c, k)) for k, *_ in METRICS)

    # ---- ignored duplicate survey submissions (optional) ----
    # {norm club: [pseudo-record, ...]} for submissions that were NOT the one kept.
    discards = {}
    if SURVEY:
        S_NUM = {'msTicket': 249, 'seasonTicket': 250, 'frontShirt': 109, 'backShirt': 131, 'sleeve': 153,
                 'tvBoard': 240, 'nonTvBoard': 242, 'mdHosp': 251, 'seasonHosp': 260, 'progAd': 248,
                 'emailDb': 276, 'optedIn': 277}
        S_TERM = {'frontTerm': 108, 'backTerm': 130, 'sleeveTerm': 152}
        S_NAME = {'fsSponsor': 89, 'bsSponsor': 111, 'slSponsor': 133}
        S_STAND = [(155, 175), (176, 196), (197, 217), (218, 238)]
        S_MATCH = ['frontShirt', 'msTicket', 'seasonTicket', 'emailDb', 'backShirt', 'sleeve']

        def numv(v):
            return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None

        def clubof(r):
            for ci in range(11, 82):
                if r[ci] not in (None, ''):
                    return str(r[ci]).strip()
            return None

        def to_record(r, club, division):
            rec = {'club': club, 'division': division, 'metrics': {}, 'chips': {}}
            for k, ci in S_NUM.items():
                rec['metrics'][k] = {'value': numv(r[ci])}
            for k, ci in S_TERM.items():
                rec['metrics'][k] = {'value': r[ci]}   # raw text ("One year", etc.)
            for k, ci in S_NAME.items():
                rec[k] = str(r[ci]).strip() if r[ci] not in (None, '') else ''
            stands = []
            for ncol, icol in S_STAND:
                nm = r[ncol]; inc = numv(r[icol])
                if (nm not in (None, '')) or inc is not None:
                    stands.append({'name': str(nm).strip() if nm not in (None, '') else '—', 'sector': '', 'income': r[icol]})
            rec['stands'] = stands
            rec['metrics']['standCount'] = {'value': float(len(stands))}
            return rec

        def completeness(rec):
            n = sum(1 for k in list(S_NUM) + list(S_TERM) if rec['metrics'].get(k, {}).get('value') not in (None, ''))
            n += sum(1 for k in S_NAME if rec.get(k))
            return n + len(rec.get('stands', []))

        def score(rec, live):
            s = 0
            for k in S_MATCH:
                a, b = rec['metrics'].get(k, {}).get('value'), mval(live, k)
                if isnum(a) and isnum(b) and abs(a - b) < 0.5:
                    s += 1
            if rec.get('fsSponsor') and norm(rec['fsSponsor']) == norm(live.get('fsSponsor')):
                s += 1
            return s

        ws_s = openpyxl.load_workbook(SURVEY, data_only=True).active
        rows = list(ws_s.iter_rows(values_only=True))[2:]
        live_by = {norm(c.get('club', '')): c for c in clubs}
        subs = {}
        for r in rows:
            cl = clubof(r)
            if cl:
                subs.setdefault(norm(cl), {'name': cl, 'rows': []})['rows'].append(r)
        for key, g in subs.items():
            if len(g['rows']) < 2:
                continue
            live = live_by.get(key, {})
            recs = [to_record(r, g['name'], live.get('division', '')) for r in g['rows']]
            kept = max(range(len(recs)), key=lambda i: score(recs[i], live)) if live else 0
            ignored = [recs[i] for i in range(len(recs)) if i != kept and completeness(recs[i]) >= 3]
            if ignored:
                discards[key] = ignored

    STRIKE = Font(strike=True, color='9A9A9A')

    wb = Workbook()

    # ================= sheet 1: All clubs =================
    ws = wb.active
    ws.title = 'All clubs'
    cols = ['Club', 'Division', 'Status']
    metric_cols = {}
    for key, h, _n, _f in METRICS:
        metric_cols[key] = len(cols)  # 0-based index into the row list
        cols.append(h)
    for _k, h in SPON:
        cols.append(h)
    for i in (1, 2, 3, 4):
        cols += ['Stand %d name' % i, 'Stand %d sector' % i, 'Stand %d income (£)' % i]
    for _k, h in CHIPS:
        cols.append(h)
    cols += ['Outlier flags', 'Data-quality flags']
    ws.append(cols)

    def base_cells(rec):
        cells = []
        for key, _h, _n, _f in METRICS:
            cells.append(money(mval(rec, key)))
        for k, _h in SPON:
            cells.append(rec.get(k, '') or '')
        sts = rec.get('stands') or []
        for i in range(4):
            sd = sts[i] if i < len(sts) else {}
            nm = sd.get('name', '') if isinstance(sd, dict) else ''
            cells += [('' if nm == '—' else nm), sd.get('sector', '') if isinstance(sd, dict) else '',
                      money(sd.get('income')) if isinstance(sd, dict) else '']
        ch = rec.get('chips') or {}
        for k, _h in CHIPS:
            cells.append(ch.get(k, '') or '')
        return cells

    for c in clubs:
        club = c.get('club', '')
        outs = outliers_for(c)
        notes = '; '.join('%s %s (%s× median)' % (HDR[k], money(v), round(r, 1) if r else '?')
                          for k, v, _med, r in outs)
        qf = quality_flags(c)
        row = [club, c.get('division', ''), 'OK' if has_data(c) else 'NO DATA SUBMITTED'] + base_cells(c) \
            + [notes, '; '.join('%s — %s' % (t, d) for t, d in qf)]
        ws.append(row)
        r = ws.max_row
        if qf:
            for ci in range(1, len(cols) + 1):
                ws.cell(r, ci).fill = RED
        for k, _v, _med, _ratio in outs:
            ws.cell(r, metric_cols[k] + 1).fill = AMBER

        # ignored duplicate submissions, struck through beneath the kept row
        for drec in discards.get(norm(club), []):
            drow = [club, c.get('division', ''), 'IGNORED — duplicate submission'] + base_cells(drec) \
                + ['', 'Superseded by the kept submission above — shown for reference']
            ws.append(drow)
            rr = ws.max_row
            for ci in range(1, len(cols) + 1):
                ws.cell(rr, ci).font = STRIKE

    # ================= sheet 2: Outliers =================
    wo = wb.create_sheet('Outliers')
    wo.append(['Club', 'Division', 'Field', 'Value', 'League median', '× median'])
    flat = []
    for c in clubs:
        for k, v, med, ratio in outliers_for(c):
            flat.append([c.get('club', ''), c.get('division', ''), HDR[k], money(v),
                         money(med), round(ratio, 1) if ratio else None])
    flat.sort(key=lambda x: (x[5] is None, -(x[5] or 0)))
    for f in flat:
        wo.append(f)

    # ================= sheet 3: Data quality =================
    wd = wb.create_sheet('Data quality')
    wd.append(['Club', 'Division', 'Issue', 'Detail'])
    qrows = []
    for c in clubs:
        for t, d in quality_flags(c):
            qrows.append([c.get('club', ''), c.get('division', ''), t, d])
    for q in qrows:
        wd.append(q)

    # ---- styling pass: header row, borders, widths, freeze ----
    for sheet in wb.worksheets:
        for ci in range(1, sheet.max_column + 1):
            cell = sheet.cell(1, ci)
            cell.fill = HEAD; cell.font = HEADF
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        sheet.freeze_panes = 'A2'
        for ci in range(1, sheet.max_column + 1):
            width = 12
            for ri in range(1, min(sheet.max_row, 80) + 1):
                v = sheet.cell(ri, ci).value
                if v is not None:
                    width = max(width, min(40, len(str(v)) + 2))
            sheet.column_dimensions[get_column_letter(ci)].width = width
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions[get_column_letter(len(cols) - 1)].width = 48  # Outlier flags
    ws.column_dimensions[get_column_letter(len(cols))].width = 60       # Data-quality flags
    wd.column_dimensions['D'].width = 80

    wb.save(OUT)
    nd = sum(1 for c in clubs if not has_data(c))
    print('Wrote %s' % OUT)
    print('  %d clubs (%d with data, %d no data submitted)' % (len(clubs), len(clubs) - nd, nd))
    print('  %d outlier values flagged across %d clubs'
          % (len(flat), len(set(f[0] for f in flat))))
    print('  %d data-quality flag(s) across %d clubs'
          % (len(qrows), len(set(q[0] for q in qrows))))


if __name__ == '__main__':
    main()
